from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.models.cultivo import Cultivo, Insumo
from app.models.plaga import DeteccionPlaga

def generar_dashboard_ejecutivo_excel(filepath, user_id):
    wb = Workbook()

    wb.remove(wb.active)

    ws_resumen = wb.create_sheet("Resumen Ejecutivo")

    ws_resumen['A1'] = 'AGROYACHAY - DASHBOARD EJECUTIVO'
    ws_resumen['A1'].font = Font(size=24, bold=True, color="FFFFFF")
    ws_resumen['A1'].fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_resumen.merge_cells('A1:F1')
    ws_resumen.row_dimensions[1].height = 40

    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws_resumen['A2'] = f'Fecha de generación: {fecha_actual}'
    ws_resumen['A2'].font = Font(italic=True)
    ws_resumen['A2'].alignment = Alignment(horizontal='center')
    ws_resumen.merge_cells('A2:F2')

    cultivos = Cultivo.query.filter_by(usuario_id=user_id).all()
    cultivos_activos = [c for c in cultivos if c.estado in ['Sembrado', 'Crecimiento', 'Floración', 'Maduración']]
    area_total = sum([float(c.area_hectareas) for c in cultivos])

    total_detecciones = 0
    if cultivos:
        from app.models import db
        total_detecciones = db.session.query(db.func.count(DeteccionPlaga.id))\
            .join(Cultivo).filter(Cultivo.usuario_id == user_id).scalar() or 0

    insumos_total = 0
    costo_total_insumos = 0
    for cultivo in cultivos:
        insumos = Insumo.query.filter_by(cultivo_id=cultivo.id).all()
        insumos_total += len(insumos)
        costo_total_insumos += sum([float(i.costo_total) for i in insumos if i.costo_total])

    ws_resumen['A4'] = 'MÉTRICAS PRINCIPALES'
    ws_resumen['A4'].font = Font(size=16, bold=True, color="2E7D32")
    ws_resumen.merge_cells('A4:B4')

    metricas = [
        ['Metrica', 'Valor', 'Estado'],
        ['Total de Cultivos', len(cultivos), 'Excelente' if len(cultivos) > 5 else 'Regular'],
        ['Cultivos Activos', len(cultivos_activos), 'Excelente' if len(cultivos_activos) > 3 else 'Regular'],
        ['Area Total (ha)', f'{area_total:.2f}', 'Excelente' if area_total > 5 else 'Regular'],
        ['Detecciones de Plagas', total_detecciones, 'Bajo' if total_detecciones < 5 else 'Alto'],
        ['Insumos Utilizados', insumos_total, 'Normal'],
        ['Inversion Total (S/.)', f'{costo_total_insumos:.2f}', 'Registrado']
    ]

    for row_idx, row_data in enumerate(metricas, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_resumen.cell(row=row_idx, column=col_idx, value=value)

            if row_idx == 5:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="43A047", end_color="43A047", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                if col_idx == 1:
                    cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")

            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 15
    ws_resumen.column_dimensions['C'].width = 20

    ws_cultivos = wb.create_sheet("Cultivos por Tipo")

    ws_cultivos['A1'] = 'DISTRIBUCIÓN DE CULTIVOS POR TIPO'
    ws_cultivos['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws_cultivos['A1'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    ws_cultivos['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_cultivos.merge_cells('A1:D1')
    ws_cultivos.row_dimensions[1].height = 35

    cultivos_por_tipo = {}
    area_por_tipo = {}
    for cultivo in cultivos:
        tipo = cultivo.tipo_cultivo
        cultivos_por_tipo[tipo] = cultivos_por_tipo.get(tipo, 0) + 1
        area_por_tipo[tipo] = area_por_tipo.get(tipo, 0) + float(cultivo.area_hectareas)

    ws_cultivos['A3'] = 'Tipo de Cultivo'
    ws_cultivos['B3'] = 'Cantidad'
    ws_cultivos['C3'] = 'Área Total (ha)'
    ws_cultivos['D3'] = 'Promedio (ha/cultivo)'

    for cell in [ws_cultivos['A3'], ws_cultivos['B3'], ws_cultivos['C3'], ws_cultivos['D3']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_num = 4
    for tipo in sorted(cultivos_por_tipo.keys()):
        cantidad = cultivos_por_tipo[tipo]
        area = area_por_tipo[tipo]
        promedio = area / cantidad if cantidad > 0 else 0

        ws_cultivos.cell(row=row_num, column=1, value=tipo).alignment = Alignment(horizontal='left')
        ws_cultivos.cell(row=row_num, column=2, value=cantidad).alignment = Alignment(horizontal='center')
        ws_cultivos.cell(row=row_num, column=3, value=round(area, 2)).alignment = Alignment(horizontal='center')
        ws_cultivos.cell(row=row_num, column=4, value=round(promedio, 2)).alignment = Alignment(horizontal='center')

        if row_num % 2 == 0:
            for col in range(1, 5):
                ws_cultivos.cell(row=row_num, column=col).fill = PatternFill(
                    start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"
                )

        row_num += 1

    ws_cultivos.column_dimensions['A'].width = 25
    ws_cultivos.column_dimensions['B'].width = 15
    ws_cultivos.column_dimensions['C'].width = 20
    ws_cultivos.column_dimensions['D'].width = 25

    if len(cultivos_por_tipo) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Cantidad de Cultivos por Tipo"
        chart.y_axis.title = 'Cantidad'
        chart.x_axis.title = 'Tipo de Cultivo'

        data = Reference(ws_cultivos, min_col=2, min_row=3, max_row=row_num-1)
        cats = Reference(ws_cultivos, min_col=1, min_row=4, max_row=row_num-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width = 20

        ws_cultivos.add_chart(chart, "F3")

        pie_chart = PieChart()
        pie_chart.title = "Distribución de Área por Tipo"

        data_pie = Reference(ws_cultivos, min_col=3, min_row=3, max_row=row_num-1)
        labels = Reference(ws_cultivos, min_col=1, min_row=4, max_row=row_num-1)
        pie_chart.add_data(data_pie, titles_from_data=True)
        pie_chart.set_categories(labels)
        pie_chart.height = 12
        pie_chart.width = 20

        ws_cultivos.add_chart(pie_chart, "F23")

    ws_financiero = wb.create_sheet("Análisis Financiero")

    ws_financiero['A1'] = 'ANÁLISIS FINANCIERO'
    ws_financiero['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws_financiero['A1'].fill = PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid")
    ws_financiero['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_financiero.merge_cells('A1:E1')
    ws_financiero.row_dimensions[1].height = 35

    ws_financiero['A3'] = 'Cultivo'
    ws_financiero['B3'] = 'Tipo'
    ws_financiero['C3'] = 'Área (ha)'
    ws_financiero['D3'] = 'Costo Insumos (S/.)'
    ws_financiero['E3'] = 'Costo/ha (S/.)'

    for cell in [ws_financiero['A3'], ws_financiero['B3'], ws_financiero['C3'], ws_financiero['D3'], ws_financiero['E3']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_num = 4
    for cultivo in cultivos:
        insumos = Insumo.query.filter_by(cultivo_id=cultivo.id).all()
        costo_cultivo = sum([float(i.costo_total) for i in insumos if i.costo_total])
        area = float(cultivo.area_hectareas)
        costo_por_ha = costo_cultivo / area if area > 0 else 0

        ws_financiero.cell(row=row_num, column=1, value=cultivo.nombre)
        ws_financiero.cell(row=row_num, column=2, value=cultivo.tipo_cultivo)
        ws_financiero.cell(row=row_num, column=3, value=round(area, 2))
        ws_financiero.cell(row=row_num, column=4, value=round(costo_cultivo, 2))
        ws_financiero.cell(row=row_num, column=5, value=round(costo_por_ha, 2))

        if row_num % 2 == 0:
            for col in range(1, 6):
                ws_financiero.cell(row=row_num, column=col).fill = PatternFill(
                    start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
                )

        row_num += 1

    ws_financiero.cell(row=row_num, column=1, value='TOTAL').font = Font(bold=True, size=12)
    ws_financiero.cell(row=row_num, column=4, value=round(costo_total_insumos, 2)).font = Font(bold=True, size=12)
    for col in range(1, 6):
        ws_financiero.cell(row=row_num, column=col).fill = PatternFill(
            start_color="FFCC80", end_color="FFCC80", fill_type="solid"
        )

    ws_financiero.column_dimensions['A'].width = 25
    ws_financiero.column_dimensions['B'].width = 20
    ws_financiero.column_dimensions['C'].width = 15
    ws_financiero.column_dimensions['D'].width = 20
    ws_financiero.column_dimensions['E'].width = 18

    ws_detalle = wb.create_sheet("Detalle de Cultivos")

    ws_detalle['A1'] = 'DETALLE COMPLETO DE CULTIVOS'
    ws_detalle['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws_detalle['A1'].fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
    ws_detalle['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_detalle.merge_cells('A1:H1')
    ws_detalle.row_dimensions[1].height = 35

    headers = ['Nombre', 'Tipo', 'Variedad', 'Área (ha)', 'Fecha Siembra', 'Estado', 'Días Desde Siembra', 'Rendimiento Esperado']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_detalle.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_num = 4
    for cultivo in cultivos:
        dias_desde_siembra = (datetime.now().date() - cultivo.fecha_siembra).days if cultivo.fecha_siembra else 0

        rendimiento_map = {
            'Maíz': 8.5, 'Trigo': 5.2, 'Arroz': 7.8, 'Papa': 25.0,
            'Tomate': 45.0, 'Cebolla': 35.0, 'Lechuga': 22.0
        }
        rendimiento_base = rendimiento_map.get(cultivo.tipo_cultivo, 10.0)
        rendimiento_esperado = round(rendimiento_base * float(cultivo.area_hectareas), 2)

        ws_detalle.cell(row=row_num, column=1, value=cultivo.nombre)
        ws_detalle.cell(row=row_num, column=2, value=cultivo.tipo_cultivo)
        ws_detalle.cell(row=row_num, column=3, value=cultivo.variedad or 'N/A')
        ws_detalle.cell(row=row_num, column=4, value=round(float(cultivo.area_hectareas), 2))
        ws_detalle.cell(row=row_num, column=5, value=cultivo.fecha_siembra.strftime('%d/%m/%Y') if cultivo.fecha_siembra else 'N/A')
        ws_detalle.cell(row=row_num, column=6, value=cultivo.estado)
        ws_detalle.cell(row=row_num, column=7, value=dias_desde_siembra)
        ws_detalle.cell(row=row_num, column=8, value=f"{rendimiento_esperado} ton")

        estado_cell = ws_detalle.cell(row=row_num, column=6)
        if cultivo.estado == 'Cosechado':
            estado_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            estado_cell.font = Font(bold=True, color="2E7D32")
        elif cultivo.estado in ['Crecimiento', 'Floración']:
            estado_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            estado_cell.font = Font(bold=True, color="F57F17")
        elif cultivo.estado == 'Sembrado':
            estado_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            estado_cell.font = Font(bold=True, color="1565C0")

        if row_num % 2 == 0:
            for col in range(1, 9):
                if col != 6:
                    ws_detalle.cell(row=row_num, column=col).fill = PatternFill(
                        start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"
                    )

        row_num += 1

    for col in range(1, 9):
        ws_detalle.column_dimensions[get_column_letter(col)].width = 18

    ws_plagas = wb.create_sheet("Análisis de Plagas")

    ws_plagas['A1'] = 'ANÁLISIS DE DETECCIÓN DE PLAGAS'
    ws_plagas['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws_plagas['A1'].fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    ws_plagas['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_plagas.merge_cells('A1:E1')
    ws_plagas.row_dimensions[1].height = 35

    ws_plagas['A3'] = 'RESUMEN GENERAL'
    ws_plagas['A3'].font = Font(size=14, bold=True, color="D32F2F")
    ws_plagas.merge_cells('A3:B3')

    ws_plagas['A4'] = 'Total de Detecciones:'
    ws_plagas['B4'] = total_detecciones
    ws_plagas['B4'].font = Font(bold=True, size=12)

    ws_plagas['A5'] = 'Cultivos Afectados:'
    from app.models import db
    cultivos_con_plagas = db.session.query(db.func.count(db.func.distinct(DeteccionPlaga.cultivo_id)))\
        .join(Cultivo).filter(Cultivo.usuario_id == user_id).scalar() or 0
    ws_plagas['B5'] = cultivos_con_plagas
    ws_plagas['B5'].font = Font(bold=True, size=12)

    ws_plagas['A6'] = 'Nivel de Riesgo:'
    if total_detecciones == 0:
        nivel_riesgo = 'BAJO - Sin plagas detectadas'
        color_riesgo = "4CAF50"
    elif total_detecciones <= 5:
        nivel_riesgo = 'MODERADO - Monitoreo regular'
        color_riesgo = "FFC107"
    else:
        nivel_riesgo = 'ALTO - Acción inmediata requerida'
        color_riesgo = "F44336"

    ws_plagas['B6'] = nivel_riesgo
    ws_plagas['B6'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_plagas['B6'].fill = PatternFill(start_color=color_riesgo, end_color=color_riesgo, fill_type="solid")
    ws_plagas.merge_cells('B6:E6')
    ws_plagas['B6'].alignment = Alignment(horizontal='center')

    ws_plagas.column_dimensions['A'].width = 25
    ws_plagas.column_dimensions['B'].width = 30

    ws_recom = wb.create_sheet("Recomendaciones")

    ws_recom['A1'] = 'RECOMENDACIONES AGRONÓMICAS'
    ws_recom['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws_recom['A1'].fill = PatternFill(start_color="00796B", end_color="00796B", fill_type="solid")
    ws_recom['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_recom.merge_cells('A1:C1')
    ws_recom.row_dimensions[1].height = 35

    recomendaciones = []

    if len(cultivos_activos) / len(cultivos) < 0.5 if cultivos else False:
        recomendaciones.append(('Productividad', 'Menos del 50% de cultivos están activos. Revisar calendario de siembra.', 'ALTA'))

    if costo_total_insumos / area_total < 500 if area_total > 0 else False:
        recomendaciones.append(('Inversión', 'Inversión por hectárea baja. Considerar fertilizantes de mejor calidad.', 'MEDIA'))

    if total_detecciones > len(cultivos) * 1.5:
        recomendaciones.append(('Plagas', 'Alto nivel de plagas. Implementar manejo integrado de plagas (MIP).', 'CRÍTICA'))

    if len(set([c.tipo_cultivo for c in cultivos])) <= 2:
        recomendaciones.append(('Diversificación', 'Poca diversificación de cultivos. Reducir riesgo con más variedades.', 'MEDIA'))

    if not recomendaciones:
        recomendaciones.append(('Operación Óptima', 'Excelente! Tu operación está funcionando correctamente.', 'BAJA'))

    ws_recom['A3'] = 'Categoría'
    ws_recom['B3'] = 'Recomendación'
    ws_recom['C3'] = 'Prioridad'
    for cell in [ws_recom['A3'], ws_recom['B3'], ws_recom['C3']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="00897B", end_color="00897B", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_num = 4
    for categoria, texto, prioridad in recomendaciones:
        ws_recom.cell(row=row_num, column=1, value=categoria).font = Font(bold=True)
        ws_recom.cell(row=row_num, column=2, value=texto)

        prioridad_cell = ws_recom.cell(row=row_num, column=3, value=prioridad)
        prioridad_cell.alignment = Alignment(horizontal='center')
        prioridad_cell.font = Font(bold=True, color="FFFFFF")

        if prioridad == 'CRÍTICA':
            prioridad_cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        elif prioridad == 'ALTA':
            prioridad_cell.fill = PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid")
        elif prioridad == 'MEDIA':
            prioridad_cell.fill = PatternFill(start_color="FBC02D", end_color="FBC02D", fill_type="solid")
        else:
            prioridad_cell.fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")

        row_num += 1

    ws_recom.column_dimensions['A'].width = 20
    ws_recom.column_dimensions['B'].width = 60
    ws_recom.column_dimensions['C'].width = 15

    wb.save(filepath)
    return True


def generar_estado_cultivos_excel(filepath, user_id, parametros):
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Cultivos"

    ws['A1'] = 'INFORME DE ESTADO DE CULTIVOS'
    ws['A1'].font = Font(size=20, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 40

    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:F2')

    headers = ['Cultivo', 'Tipo', 'Área (ha)', 'Estado', 'Fecha Siembra', 'Días']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="43A047", end_color="43A047", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    cultivos = Cultivo.query.filter_by(usuario_id=user_id).all()
    for row_idx, cultivo in enumerate(cultivos, start=5):
        dias = (datetime.now().date() - cultivo.fecha_siembra).days if cultivo.fecha_siembra else 0

        ws.cell(row=row_idx, column=1, value=cultivo.nombre)
        ws.cell(row=row_idx, column=2, value=cultivo.tipo_cultivo)
        ws.cell(row=row_idx, column=3, value=round(float(cultivo.area_hectareas), 2))
        ws.cell(row=row_idx, column=4, value=cultivo.estado)
        ws.cell(row=row_idx, column=5, value=cultivo.fecha_siembra.strftime('%d/%m/%Y') if cultivo.fecha_siembra else 'N/A')
        ws.cell(row=row_idx, column=6, value=dias)

        if row_idx % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="F1F8E9", end_color="F1F8E9", fill_type="solid"
                )

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(filepath)
    return True


def generar_analisis_financiero_excel(filepath, user_id, parametros):
    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis Financiero"

    ws['A1'] = 'ANÁLISIS FINANCIERO DETALLADO'
    ws['A1'].font = Font(size=20, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 40

    wb.save(filepath)
    return True


def generar_impacto_climatico_excel(filepath, user_id, parametros):
    wb = Workbook()
    ws = wb.active
    ws.title = "Impacto Climático"

    wb.save(filepath)
    return True
